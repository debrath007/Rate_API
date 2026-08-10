"""Tests for check_compliance.py.

Fixtures are built on the fly with openpyxl into tmp_path, so there is no binary
test file to keep in sync and the suite runs anywhere the script does.

    pytest test_check_compliance.py
"""

from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).parent))

import check_compliance as cc  # noqa: E402

SCRIPT = Path(__file__).parent / "check_compliance.py"
RULES = Path(__file__).parent.parent / "rules.json"

FULL_HEADERS = ["AccountId", "RateType", "Scenario", "RateBasis", "Index", "Margin",
                "Rate", "FloorRate", "CeilingRate", "OverrideCode", "OverrideRate"]

MINIMAL_HEADERS = ["AccountId", "RateType", "Rate", "OverrideCode", "OverrideRate"]

PCT = {"Index", "Margin", "Rate", "FloorRate", "CeilingRate", "OverrideRate"}


def write_workbook(path, rows, headers=None, as_percent=True):
    """rows: dicts keyed by header name. Missing keys are left blank."""
    headers = headers or FULL_HEADERS
    wb = Workbook()
    ws = wb.active
    ws.title = "APR Report"
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    for r, row in enumerate(rows, start=2):
        for col, header in enumerate(headers, start=1):
            value = row.get(header)
            if value is None:
                continue
            if header in PCT and as_percent:
                cell = ws.cell(row=r, column=col, value=value / 100.0)
                cell.number_format = "0.00%"
            else:
                ws.cell(row=r, column=col, value=value)
    wb.save(path)
    return path


def row(account="ACC1", rate_type="01", **kwargs):
    """A compliant variable-rate row; override individual fields per test."""
    base = {"AccountId": account, "RateType": rate_type, "Scenario": "TEST",
            "RateBasis": "VARIABLE", "Index": 8.25, "Margin": 10.74, "Rate": 18.99,
            "FloorRate": 9.99, "CeilingRate": 29.99}
    base.update({k: v for k, v in kwargs.items()})
    return base


def check(path, sheet=None, ruleset=None):
    return cc.build_report(cc.load_rows(str(path), sheet), ruleset)


def rules(report, rule):
    return [v for v in report["violations"] if v["rule"] == rule]


# --- OVERRIDE_CODE_CEILING ---------------------------------------------------

def test_clean_sheet_is_compliant(tmp_path):
    path = write_workbook(tmp_path / "clean.xlsx", [
        row("ACC1", "01"),
        row("ACC1", "02", OverrideCode="SCRA", OverrideRate=6.00),
        row("ACC2", "03", OverrideCode="CMA", OverrideRate=12.00),
        row("ACC2", "04", RateBasis="FIXED", Index=None, Margin=None, Rate=19.99),
    ])
    report = check(path)

    assert report["status"] == "compliant"
    assert report["violations"] == []
    assert report["rowsChecked"] == 4
    assert report["accountsChecked"] == 2


def test_scra_at_the_cap_is_not_flagged(tmp_path):
    path = write_workbook(tmp_path / "at.xlsx",
                          [row(OverrideCode="SCRA", OverrideRate=6.00)])
    assert rules(check(path), cc.RULE_CODE_CEILING) == []


def test_scra_below_the_cap_is_compliant(tmp_path):
    # APR-099: the cap is a ceiling, not a fixed rate. A customer whose formula lands
    # under it keeps the lower rate, and that is the correct outcome.
    path = write_workbook(tmp_path / "below.xlsx", [
        row(Index=2.00, Margin=2.50, Rate=4.50, OverrideCode="SCRA", OverrideRate=6.00)])
    assert check(path)["violations"] == []


def test_scra_above_the_cap_is_critical(tmp_path):
    path = write_workbook(tmp_path / "above.xlsx",
                          [row(OverrideCode="SCRA", OverrideRate=7.50)])

    found = rules(check(path), cc.RULE_CODE_CEILING)

    assert len(found) == 1
    assert found[0]["severity"] == cc.SEVERITY_CRITICAL
    assert found[0]["expected"] == 6.00
    assert found[0]["actual"] == 7.50
    assert "above the 6.00% cap" in found[0]["message"]


def test_mla_has_its_own_higher_ceiling(tmp_path):
    path = write_workbook(tmp_path / "mla.xlsx", [
        row(Index=8.25, Margin=20.00, Rate=28.25, CeilingRate=36.00,
            OverrideCode="MLA", OverrideRate=36.00)])
    assert rules(check(path), cc.RULE_CODE_CEILING) == []


def test_most_protective_ceiling_wins_when_several_codes_apply(tmp_path):
    # APR-162: SCRA's 6% is stricter than MLA's 36%, so 6% governs.
    path = write_workbook(tmp_path / "both.xlsx", [
        row(CeilingRate=36.00, OverrideCode="SCRA,MLA", OverrideRate=10.00)])

    found = rules(check(path), cc.RULE_CODE_CEILING)

    assert len(found) == 1
    assert found[0]["expected"] == 6.00
    assert "SCRA" in found[0]["message"]


def test_unmapped_override_code_has_no_ceiling(tmp_path):
    path = write_workbook(tmp_path / "cma.xlsx",
                          [row(OverrideCode="CMA", OverrideRate=18.00)])
    assert rules(check(path), cc.RULE_CODE_CEILING) == []


# --- MAX_APR_CAP -------------------------------------------------------------

def test_rate_above_the_row_ceiling_is_critical(tmp_path):
    path = write_workbook(tmp_path / "cap.xlsx",
                          [row(Index=8.25, Margin=23.24, Rate=31.49)])

    found = rules(check(path), cc.RULE_MAX_APR)

    assert len(found) == 1
    assert found[0]["severity"] == cc.SEVERITY_CRITICAL
    assert found[0]["expected"] == 29.99
    assert found[0]["actual"] == 31.49
    assert "maximum APR" in found[0]["message"]


def test_rate_exactly_at_the_ceiling_is_not_flagged(tmp_path):
    path = write_workbook(tmp_path / "at.xlsx",
                          [row(Index=8.25, Margin=21.74, Rate=29.99)])
    assert check(path)["violations"] == []


def test_a_hair_over_the_ceiling_is_within_tolerance(tmp_path):
    path = write_workbook(tmp_path / "eps.xlsx",
                          [row(Index=8.25, Margin=21.742, Rate=29.992)])
    assert rules(check(path), cc.RULE_MAX_APR) == []


def test_ceiling_is_read_per_row_not_from_a_global_constant(tmp_path):
    # 24.00% is fine under the default 29.99 but breaches this row's own 19.99 ceiling.
    path = write_workbook(tmp_path / "own.xlsx", [
        row(Index=8.25, Margin=15.75, Rate=24.00, FloorRate=5.00, CeilingRate=19.99)])

    found = rules(check(path), cc.RULE_MAX_APR)

    assert len(found) == 1
    assert found[0]["expected"] == 19.99


def test_high_normal_rate_masked_by_lower_override_is_not_flagged(tmp_path):
    path = write_workbook(tmp_path / "masked.xlsx", [
        row(Index=8.25, Margin=26.75, Rate=35.00, OverrideCode="CMA", OverrideRate=20.00)])
    assert rules(check(path), cc.RULE_MAX_APR) == []


def test_override_above_the_ceiling_is_flagged_when_it_is_the_lower_rate(tmp_path):
    path = write_workbook(tmp_path / "ovcap.xlsx", [
        row(Index=8.25, Margin=31.75, Rate=40.00, OverrideCode="CMA", OverrideRate=32.00)])

    found = rules(check(path), cc.RULE_MAX_APR)

    assert len(found) == 1
    assert found[0]["actual"] == 32.00


# --- RATE_FLOOR --------------------------------------------------------------

def test_rate_below_the_floor_is_high(tmp_path):
    path = write_workbook(tmp_path / "floor.xlsx",
                          [row(Index=0.50, Margin=4.49, Rate=4.99)])

    found = rules(check(path), cc.RULE_FLOOR)

    assert len(found) == 1
    assert found[0]["severity"] == cc.SEVERITY_HIGH
    assert found[0]["expected"] == 9.99
    assert found[0]["actual"] == 4.99


def test_rate_exactly_at_the_floor_is_not_flagged(tmp_path):
    path = write_workbook(tmp_path / "atfloor.xlsx",
                          [row(Index=0.50, Margin=9.49, Rate=9.99)])
    assert rules(check(path), cc.RULE_FLOOR) == []


def test_floor_does_not_apply_to_overridden_rows(tmp_path):
    # An override is a legitimate reason to sit below the floor; applying the floor
    # here would flag every SCRA and goodwill account.
    path = write_workbook(tmp_path / "ovfloor.xlsx",
                          [row(OverrideCode="SCRA", OverrideRate=6.00)])
    assert rules(check(path), cc.RULE_FLOOR) == []


# --- LOWER_RATE_WINS ---------------------------------------------------------

def test_override_code_without_a_rate_is_reported(tmp_path):
    path = write_workbook(tmp_path / "norate.xlsx", [row(OverrideCode="CMA")])

    found = rules(check(path), cc.RULE_LOWER_WINS)

    assert len(found) == 1
    assert found[0]["severity"] == cc.SEVERITY_HIGH
    assert found[0]["expected"] is None
    assert found[0]["actual"] == 18.99
    assert "no override rate" in found[0]["message"]


def test_override_higher_than_normal_rate_is_not_a_violation(tmp_path):
    path = write_workbook(tmp_path / "higher.xlsx", [
        row(Index=8.00, Margin=0.00, Rate=8.00, FloorRate=5.00,
            OverrideCode="CMA", OverrideRate=9.00)])
    assert check(path)["violations"] == []


# --- RATE_MATCHES_FORMULA ----------------------------------------------------

def test_stored_rate_that_does_not_match_index_plus_margin_is_flagged(tmp_path):
    path = write_workbook(tmp_path / "drift.xlsx",
                          [row(Index=8.25, Margin=14.74, Rate=23.05)])

    found = rules(check(path), cc.RULE_FORMULA)

    assert len(found) == 1
    assert found[0]["severity"] == cc.SEVERITY_HIGH
    assert found[0]["expected"] == 22.99
    assert found[0]["actual"] == 23.05


def test_formula_is_not_applied_to_fixed_rate_rows(tmp_path):
    path = write_workbook(tmp_path / "fixed.xlsx", [
        row(RateBasis="FIXED", Index=None, Margin=None, Rate=19.99)])
    assert rules(check(path), cc.RULE_FORMULA) == []


def test_formula_checks_the_normal_rate_not_the_overridden_one(tmp_path):
    # The override changes what is charged, not how the underlying rate is derived.
    path = write_workbook(tmp_path / "ovformula.xlsx", [
        row(Index=8.25, Margin=10.74, Rate=18.99, OverrideCode="CMA", OverrideRate=7.00)])
    assert rules(check(path), cc.RULE_FORMULA) == []


# --- RATE_SANITY -------------------------------------------------------------

def test_negative_rate_is_critical(tmp_path):
    path = write_workbook(tmp_path / "neg.xlsx",
                          [row(Index=-5.00, Margin=1.00, Rate=-4.00)])

    found = rules(check(path), cc.RULE_SANITY)

    assert len(found) == 1
    assert found[0]["severity"] == cc.SEVERITY_CRITICAL
    assert found[0]["actual"] == -4.00


def test_absurd_rate_is_critical(tmp_path):
    path = write_workbook(tmp_path / "absurd.xlsx",
                          [row(Index=100.00, Margin=50.00, Rate=150.00)])

    found = rules(check(path), cc.RULE_SANITY)

    assert len(found) == 1
    assert "sanity ceiling" in found[0]["message"]


# --- Report shape ------------------------------------------------------------

def test_findings_carry_the_scenario_name(tmp_path):
    path = write_workbook(tmp_path / "scen.xlsx", [
        row(Scenario="SCRA_BREACH", OverrideCode="SCRA", OverrideRate=7.00)])

    found = rules(check(path), cc.RULE_CODE_CEILING)

    assert found[0]["scenario"] == "SCRA_BREACH"


def test_rules_are_evaluated_in_file_order(tmp_path):
    # A SCRA row with no override rate falls back to a normal rate over the ceiling,
    # tripping three rules at once. Order is part of the report contract.
    path = write_workbook(tmp_path / "order.xlsx", [
        row(Index=8.25, Margin=24.75, Rate=33.00, OverrideCode="SCRA")])

    report = check(path)

    assert [v["rule"] for v in report["violations"]] == [
        cc.RULE_CODE_CEILING, cc.RULE_MAX_APR, cc.RULE_LOWER_WINS]
    assert report["criticalCount"] == 2
    assert report["highCount"] == 1


def test_report_echoes_the_rules_that_were_applied(tmp_path):
    path = write_workbook(tmp_path / "meta.xlsx", [row()])

    report = check(path)

    # Asserted against the rule file rather than a literal: pinning the version here
    # means every policy release breaks a test that is not about versioning.
    ruleset = json.loads(RULES.read_text(encoding="utf-8"))
    enabled = [r["id"] for r in ruleset["rules"] if r.get("enabled", True)]

    assert report["rulesVersion"] == ruleset["version"]
    assert cc.RULE_CODE_CEILING in report["rulesApplied"]
    assert report["rulesApplied"] == enabled


# --- The rule file drives behaviour ------------------------------------------

def test_disabling_a_rule_silences_it(tmp_path):
    ruleset = json.loads(RULES.read_text(encoding="utf-8"))
    for rule in ruleset["rules"]:
        if rule["id"] == cc.RULE_CODE_CEILING:
            rule["enabled"] = False

    path = write_workbook(tmp_path / "off.xlsx",
                          [row(OverrideCode="SCRA", OverrideRate=7.50)])

    report = check(path, ruleset=ruleset)

    assert rules(report, cc.RULE_CODE_CEILING) == []
    assert cc.RULE_CODE_CEILING not in report["rulesApplied"]


def test_changing_a_threshold_changes_the_verdict(tmp_path):
    ruleset = json.loads(RULES.read_text(encoding="utf-8"))
    for rule in ruleset["rules"]:
        if rule["id"] == cc.RULE_CODE_CEILING:
            rule["params"]["codeCeilings"]["SCRA"] = 8.00

    path = write_workbook(tmp_path / "thresh.xlsx",
                          [row(OverrideCode="SCRA", OverrideRate=7.50)])

    assert rules(check(path, ruleset=ruleset), cc.RULE_CODE_CEILING) == []


def test_missing_rule_file_raises(tmp_path):
    with pytest.raises(cc.CheckError) as exc:
        cc.load_rules(tmp_path / "nope.json")
    assert "not found" in str(exc.value)


def test_malformed_rule_file_raises(tmp_path):
    bad = tmp_path / "bad.json"
    bad.write_text("{not json", encoding="utf-8")

    with pytest.raises(cc.CheckError) as exc:
        cc.load_rules(bad)
    assert "could not parse" in str(exc.value)


def test_unknown_check_name_lists_the_valid_ones(tmp_path):
    unknown = tmp_path / "unknown.json"
    unknown.write_text(json.dumps({"rules": [
        {"id": "X", "check": "nope", "severity": {}, "messages": {}}]}), encoding="utf-8")

    with pytest.raises(cc.CheckError) as exc:
        cc.load_rules(unknown)
    assert "Known checks" in str(exc.value)


# --- Portability -------------------------------------------------------------

def test_a_minimal_sheet_still_works(tmp_path):
    # Only the required columns: rules needing the optional ones skip rather than fail.
    path = write_workbook(tmp_path / "min.xlsx", [
        {"AccountId": "ACC1", "RateType": "01", "Rate": 31.49},
    ], headers=MINIMAL_HEADERS)

    report = check(path)

    assert rules(report, cc.RULE_MAX_APR)  # falls back to defaultMaxRate
    assert rules(report, cc.RULE_FLOOR) == []  # no FloorRate column
    assert rules(report, cc.RULE_FORMULA) == []  # no RateBasis column


def test_reordered_columns_still_work(tmp_path):
    headers = ["OverrideRate", "AccountId", "OverrideCode", "RateType", "Rate"]
    path = write_workbook(tmp_path / "reordered.xlsx", [
        {"AccountId": "ACC1", "RateType": "01", "Rate": 24.77,
         "OverrideCode": "SCRA", "OverrideRate": 7.50},
    ], headers=headers)

    found = rules(check(path), cc.RULE_CODE_CEILING)

    assert len(found) == 1
    assert found[0]["actual"] == 7.50


def test_header_matching_ignores_case_and_spacing(tmp_path):
    headers = ["account id", "RATE TYPE", "Rate", "Override_Code", "override rate"]
    path = write_workbook(tmp_path / "headers.xlsx", [
        {"AccountId": "ACC1", "RateType": "01", "Rate": 24.77,
         "OverrideCode": "SCRA", "OverrideRate": 7.50},
    ], headers=headers)
    # write_workbook keys off header text, so remap onto the odd spellings
    wb = Workbook()
    ws = wb.active
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    ws.cell(row=2, column=1, value="ACC1")
    ws.cell(row=2, column=2, value="01")
    cell = ws.cell(row=2, column=3, value=0.2477)
    cell.number_format = "0.00%"
    ws.cell(row=2, column=4, value="SCRA")
    cell = ws.cell(row=2, column=5, value=0.075)
    cell.number_format = "0.00%"
    wb.save(path)

    assert len(rules(check(path), cc.RULE_CODE_CEILING)) == 1


def test_missing_required_column_fails_loudly(tmp_path):
    path = write_workbook(tmp_path / "missing.xlsx", [
        {"AccountId": "ACC1", "RateType": "01", "Rate": 10.0, "OverrideCode": "SCRA"},
    ], headers=["AccountId", "RateType", "Rate", "OverrideCode", "Notes"])

    with pytest.raises(cc.CheckError) as exc:
        check(path)

    assert "overrideRate" in str(exc.value)
    assert "Notes" in str(exc.value)  # tells you what it did find


def test_plain_number_rate_column_is_read_as_percent(tmp_path):
    path = write_workbook(tmp_path / "plain.xlsx", [
        {"AccountId": "ACC1", "RateType": "01", "Rate": 24.77,
         "OverrideCode": "SCRA", "OverrideRate": 7.50},
    ], headers=MINIMAL_HEADERS, as_percent=False)

    found = rules(check(path), cc.RULE_CODE_CEILING)

    assert len(found) == 1
    assert found[0]["actual"] == 7.50


def test_trailing_blank_rows_are_skipped(tmp_path):
    path = write_workbook(tmp_path / "blanks.xlsx", [row(), {}])
    assert check(path)["rowsChecked"] == 1


def test_named_sheet_can_be_selected(tmp_path):
    wb = Workbook()
    wb.active.title = "Empty"
    wb.active["A1"] = "AccountId"
    target = wb.create_sheet("Rates")
    for col, header in enumerate(MINIMAL_HEADERS, start=1):
        target.cell(row=1, column=col, value=header)
    target.cell(row=2, column=1, value="ACC1")
    target.cell(row=2, column=2, value="01")
    target.cell(row=2, column=3, value=10.0)
    path = tmp_path / "sheets.xlsx"
    wb.save(path)

    assert check(path, sheet="Rates")["rowsChecked"] == 1


# --- CLI contract ------------------------------------------------------------

def run_cli(*args):
    return subprocess.run([sys.executable, str(SCRIPT), *args],
                          capture_output=True, text=True)


def test_cli_exit_code_zero_when_compliant(tmp_path):
    path = write_workbook(tmp_path / "c.xlsx", [row()])

    result = run_cli("--file", str(path))

    assert result.returncode == 0
    assert json.loads(result.stdout)["status"] == "compliant"


def test_cli_exit_code_one_when_violations_found(tmp_path):
    path = write_workbook(tmp_path / "v.xlsx",
                          [row(OverrideCode="SCRA", OverrideRate=7.50)])

    result = run_cli("--file", str(path))

    assert result.returncode == 1
    assert json.loads(result.stdout)["criticalCount"] == 1


def test_cli_exit_code_two_when_it_cannot_run(tmp_path):
    result = run_cli("--file", str(tmp_path / "nope.xlsx"))

    assert result.returncode == 2
    assert "error:" in result.stderr
    assert result.stdout == ""  # nothing parseable on stdout when the run failed


def test_cli_exit_code_two_when_the_rule_file_is_missing(tmp_path):
    path = write_workbook(tmp_path / "r.xlsx", [row()])

    result = run_cli("--file", str(path), "--rules", str(tmp_path / "nope.json"))

    assert result.returncode == 2
    assert "rule file not found" in result.stderr


def test_cli_text_format_is_human_readable(tmp_path):
    path = write_workbook(tmp_path / "t.xlsx",
                          [row(OverrideCode="SCRA", OverrideRate=7.50)])

    result = run_cli("--file", str(path), "--format", "text")

    assert result.returncode == 1
    assert "NON-COMPLIANT" in result.stdout
    assert "OVERRIDE_CODE_CEILING" in result.stdout
    assert not result.stdout.lstrip().startswith("{")


# --- Configuration rules -----------------------------------------------------
#
# These need the config columns, so they use a wider header set. The narrower
# fixtures above deliberately omit them, which is what proves these rules stay
# silent on a sheet that cannot answer them.

CONFIG_HEADERS = FULL_HEADERS + [
    "ProductCode", "OriginationDate", "OverrideExpiry",
    "ProtectionStart", "ProtectionEnd",
    "DayCountBasis", "CompoundingFrequency", "RoundingRule"]


def config_row(**kwargs):
    base = row()
    base.update({"ProductCode": "CORE", "OriginationDate": "2023-01-01",
                 "DayCountBasis": 365, "CompoundingFrequency": "DAILY",
                 "RoundingRule": "HALF_UP"})
    base.update(kwargs)
    return base


def config_check(tmp_path, name, rows_):
    return check(write_workbook(tmp_path / name, rows_, headers=CONFIG_HEADERS))


def test_config_row_is_clean(tmp_path):
    assert config_check(tmp_path, "ok.xlsx", [config_row()])["violations"] == []


def test_day_count_other_than_policy_is_critical(tmp_path):
    report = config_check(tmp_path, "dc.xlsx", [config_row(DayCountBasis=360)])

    found = rules(report, cc.RULE_DAY_COUNT)
    assert len(found) == 1
    assert found[0]["severity"] == cc.SEVERITY_CRITICAL
    assert "360" in found[0]["message"]


def test_compounding_other_than_policy_is_critical(tmp_path):
    found = rules(config_check(tmp_path, "cf.xlsx",
                               [config_row(CompoundingFrequency="MONTHLY")]),
                  cc.RULE_COMPOUNDING)
    assert len(found) == 1
    assert found[0]["severity"] == cc.SEVERITY_CRITICAL


def test_rounding_other_than_policy_is_high(tmp_path):
    found = rules(config_check(tmp_path, "rr.xlsx",
                               [config_row(RoundingRule="HALF_EVEN")]),
                  cc.RULE_ROUNDING)
    assert len(found) == 1
    assert found[0]["severity"] == cc.SEVERITY_HIGH


def test_config_rules_are_silent_when_the_columns_are_absent(tmp_path):
    # The narrow fixture has no config columns at all, so these rules must not fire.
    report = check(write_workbook(tmp_path / "narrow.xlsx", [row()]))
    assert rules(report, cc.RULE_DAY_COUNT) == []
    assert rules(report, cc.RULE_COMPOUNDING) == []
    assert rules(report, cc.RULE_ROUNDING) == []


# --- OVERRIDE_NOT_EXPIRED ----------------------------------------------------

def test_override_past_its_expiry_is_reported(tmp_path):
    found = rules(config_check(tmp_path, "exp.xlsx", [config_row(
        OverrideCode="CMA", OverrideRate=12.00, OverrideExpiry="2025-01-01")]),
        cc.RULE_OVERRIDE_EXPIRY)

    assert len(found) == 1
    assert found[0]["severity"] == cc.SEVERITY_HIGH
    assert "2025-01-01" in found[0]["message"]


def test_override_expiring_in_the_future_is_clean(tmp_path):
    assert rules(config_check(tmp_path, "future.xlsx", [config_row(
        OverrideCode="CMA", OverrideRate=12.00, OverrideExpiry="2099-01-01")]),
        cc.RULE_OVERRIDE_EXPIRY) == []


def test_expiry_without_an_override_is_not_reported(tmp_path):
    # A stale date on a row carrying no override is not charging anyone anything.
    assert rules(config_check(tmp_path, "stray.xlsx",
                              [config_row(OverrideExpiry="2025-01-01")]),
                 cc.RULE_OVERRIDE_EXPIRY) == []


# --- PROTECTION_DATES_VALID --------------------------------------------------

def test_protected_code_without_a_start_date_is_reported(tmp_path):
    found = rules(config_check(tmp_path, "nostart.xlsx", [config_row(
        OverrideCode="SCRA", OverrideRate=6.00)]), cc.RULE_PROTECTION_DATES)

    assert len(found) == 1
    assert found[0]["severity"] == cc.SEVERITY_HIGH
    assert "no protection start" in found[0]["message"]


def test_protection_that_has_ended_but_is_still_applied_is_reported(tmp_path):
    found = rules(config_check(tmp_path, "ended.xlsx", [config_row(
        OverrideCode="SCRA", OverrideRate=6.00,
        ProtectionStart="2024-01-01", ProtectionEnd="2025-06-30")]),
        cc.RULE_PROTECTION_DATES)

    assert len(found) == 1
    assert "2025-06-30" in found[0]["message"]


def test_active_protection_is_clean(tmp_path):
    assert rules(config_check(tmp_path, "active.xlsx", [config_row(
        OverrideCode="SCRA", OverrideRate=6.00, ProtectionStart="2024-01-01")]),
        cc.RULE_PROTECTION_DATES) == []


def test_unprotected_override_code_needs_no_protection_dates(tmp_path):
    # CMA is a goodwill discount, not a statutory protection.
    assert rules(config_check(tmp_path, "cma.xlsx", [config_row(
        OverrideCode="CMA", OverrideRate=12.00)]), cc.RULE_PROTECTION_DATES) == []


# --- PRE_SERVICE_DEBT_SCOPE --------------------------------------------------

def test_statutory_cap_on_post_activation_debt_is_reported(tmp_path):
    found = rules(config_check(tmp_path, "post.xlsx", [config_row(
        OverrideCode="SCRA", OverrideRate=6.00,
        OriginationDate="2025-03-01", ProtectionStart="2024-01-01")]),
        cc.RULE_PRE_SERVICE)

    assert len(found) == 1
    assert found[0]["severity"] == cc.SEVERITY_HIGH


def test_statutory_cap_on_pre_service_debt_is_clean(tmp_path):
    assert rules(config_check(tmp_path, "pre.xlsx", [config_row(
        OverrideCode="SCRA", OverrideRate=6.00,
        OriginationDate="2023-01-01", ProtectionStart="2024-01-01")]),
        cc.RULE_PRE_SERVICE) == []


# --- BOUNDS_WITHIN_PRODUCT_RANGE ---------------------------------------------

def test_ceiling_above_the_product_range_is_reported(tmp_path):
    found = rules(config_check(tmp_path, "wide.xlsx",
                               [config_row(CeilingRate=44.00)]), cc.RULE_PRODUCT_RANGE)

    assert len(found) == 1
    assert found[0]["severity"] == cc.SEVERITY_HIGH
    assert "CORE" in found[0]["message"]


def test_unknown_product_code_is_reported(tmp_path):
    found = rules(config_check(tmp_path, "mystery.xlsx",
                               [config_row(ProductCode="MYSTERY")]), cc.RULE_PRODUCT_RANGE)

    assert len(found) == 1
    assert "MYSTERY" in found[0]["message"]


def test_bounds_inside_the_product_range_are_clean(tmp_path):
    assert rules(config_check(tmp_path, "inside.xlsx", [config_row(
        ProductCode="PRIVATE", CeilingRate=36.00)]), cc.RULE_PRODUCT_RANGE) == []
