#!/usr/bin/env python3
"""Check an APR rate spreadsheet against company rate policy.

Standalone: the only third-party dependency is openpyxl. Copy this script (or the
whole skill folder it lives in) anywhere and point it at a workbook.

    python check_compliance.py --file APR_Report.xlsx
    python check_compliance.py --file APR_Report.xlsx --format text
    python check_compliance.py --file APR_Report.xlsx --rules my_rules.json

Exit codes:
    0  compliant
    1  violations found
    2  could not run (missing file, missing column, unreadable sheet, bad rule file)

The 0/1 split makes this usable directly as a CI gate.

The policy itself is documented in ../SKILL.md; this script is its executable form.
Thresholds, severities and messages are not hardcoded here -- they are read from
../rules.json, so the policy can be retuned without editing this file. This module
supplies the evaluation logic; that file supplies everything the policy decides.
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - environment problem, not logic
    print("error: openpyxl is required. Install it with: pip install openpyxl",
          file=sys.stderr)
    sys.exit(2)


# --- Rule file ---------------------------------------------------------------

# The rules live beside the skill's SKILL.md, one level up from scripts/, so that
# the policy sits at the top of the bundle rather than buried in the code folder.
DEFAULT_RULES_FILE = Path(__file__).resolve().parent.parent / "rules.json"

# Fallback tolerance, used only when the rule file omits "epsilon".
EPSILON = 0.005

# Rule ids, for callers that want to filter findings without hardcoding strings.
# These name the rules shipped in rules.json; a custom rule file may define others.
RULE_CODE_CEILING = "OVERRIDE_CODE_CEILING"
RULE_MAX_APR = "MAX_APR_CAP"
RULE_FLOOR = "RATE_FLOOR"
RULE_LOWER_WINS = "LOWER_RATE_WINS"
RULE_FORMULA = "RATE_MATCHES_FORMULA"
RULE_SANITY = "RATE_SANITY"
RULE_DAY_COUNT = "CONSISTENT_DAY_COUNT"
RULE_COMPOUNDING = "CONSISTENT_COMPOUNDING"
RULE_ROUNDING = "CONSISTENT_ROUNDING"
RULE_OVERRIDE_EXPIRY = "OVERRIDE_NOT_EXPIRED"
RULE_PROTECTION_DATES = "PROTECTION_DATES_VALID"
RULE_PRE_SERVICE = "PRE_SERVICE_DEBT_SCOPE"
RULE_PRODUCT_RANGE = "BOUNDS_WITHIN_PRODUCT_RANGE"

SEVERITY_CRITICAL = "CRITICAL"
SEVERITY_HIGH = "HIGH"

EXIT_COMPLIANT = 0
EXIT_VIOLATIONS = 1
EXIT_ERROR = 2


# --- Column handling ---------------------------------------------------------

# Logical name -> accepted header spellings (normalised: lowercased, non-alphanumerics
# stripped). Header order in the sheet does not matter.
COLUMN_ALIASES = {
    "accountId": ("accountid", "account", "accountnumber", "acctid"),
    "rateType": ("ratetype", "type", "ratecode"),
    "rate": ("rate", "normalrate", "apr", "normalapr"),
    "overrideCode": ("overridecode", "override", "code"),
    "overrideRate": ("overriderate", "overrideapr"),
}

REQUIRED_COLUMNS = tuple(COLUMN_ALIASES)

# Columns a sheet may carry to unlock further rules. A rule whose inputs are missing
# is skipped for that row rather than failing the run, so the same checker still
# works against a minimal sheet that only has the required columns above.
OPTIONAL_ALIASES = {
    "scenario": ("scenario", "case", "testcase"),
    "rateBasis": ("ratebasis", "basis", "ratetypebasis", "variableorfixed"),
    "index": ("index", "indexrate", "primerate", "benchmark"),
    "margin": ("margin", "spread"),
    "floorRate": ("floorrate", "floor", "minrate", "aprfloor"),
    "ceilingRate": ("ceilingrate", "ceiling", "maxrate", "aprceiling"),
    "productCode": ("productcode", "product", "producttype"),
    "originationDate": ("originationdate", "opened", "openeddate", "origination"),
    "overrideExpiry": ("overrideexpiry", "overrideexpires", "overrideenddate"),
    "protectionStart": ("protectionstart", "activedutystart", "protectionstartdate"),
    "protectionEnd": ("protectionend", "activedutyend", "protectionenddate"),
    "dayCountBasis": ("daycountbasis", "daycount", "divisor"),
    "compoundingFrequency": ("compoundingfrequency", "compounding"),
    "roundingRule": ("roundingrule", "rounding"),
}

# Columns holding an ISO date rather than a rate; read as text and compared
# lexically, which is ordering-correct for ISO-8601 and avoids Excel serial dates.
DATE_COLUMNS = ("originationDate", "overrideExpiry", "protectionStart", "protectionEnd")


class CheckError(Exception):
    """Something made the check impossible to run. Distinct from a violation."""


def load_rules(path=None) -> dict:
    """Read and validate the rule file.

    A malformed rule file is a CheckError, not a violation: the run could not be
    performed, so it exits 2 rather than reporting a clean sheet. Silently falling
    back to built-in defaults would be worse -- an edit with a typo in it would
    look like it had taken effect.
    """
    rules_path = Path(path) if path else DEFAULT_RULES_FILE
    try:
        with open(rules_path, encoding="utf-8") as handle:
            document = json.load(handle)
    except FileNotFoundError as exc:
        raise CheckError(f"rule file not found: {rules_path}") from exc
    except json.JSONDecodeError as exc:
        raise CheckError(f"could not parse rule file {rules_path}: {exc}") from exc
    except OSError as exc:
        raise CheckError(f"could not read rule file {rules_path}: {exc}") from exc

    if not isinstance(document, dict) or not isinstance(document.get("rules"), list):
        raise CheckError(f"rule file {rules_path} must be an object with a 'rules' array")

    seen = set()
    for index, rule in enumerate(document["rules"]):
        where = f"rule {index} in {rules_path}"
        if not isinstance(rule, dict):
            raise CheckError(f"{where} is not an object")

        rule_id = rule.get("id")
        if not rule_id:
            raise CheckError(f"{where} has no 'id'")
        if rule_id in seen:
            raise CheckError(f"duplicate rule id {rule_id!r} in {rules_path}")
        seen.add(rule_id)

        check = rule.get("check")
        if check not in CHECKS:
            raise CheckError(
                f"rule {rule_id!r} names unknown check {check!r}. "
                f"Known checks: {', '.join(sorted(CHECKS))}")

        for section in ("severity", "messages"):
            if not isinstance(rule.get(section), dict):
                raise CheckError(f"rule {rule_id!r} has no '{section}' object")
        rule.setdefault("params", {})
        # Evaluators receive a rule, not the whole document, so the shared as-of
        # date is pushed down here. A rule may still pin its own.
        if document.get("asOfDate"):
            rule["params"].setdefault("asOfDate", document["asOfDate"])

    return document


def active_rules(ruleset: dict) -> list:
    """Rules to evaluate, in file order -- which is the order findings appear in."""
    return [rule for rule in ruleset["rules"] if rule.get("enabled", True)]


def _severity(rule, key):
    try:
        return rule["severity"][key]
    except KeyError as exc:
        raise CheckError(
            f"rule {rule['id']!r} is missing severity {key!r}") from exc


def _message(rule, key, **values):
    try:
        return rule["messages"][key].format(**values)
    except KeyError as exc:
        raise CheckError(
            f"rule {rule['id']!r} has a bad message template for {key!r}: "
            f"unknown placeholder {exc}") from exc


def _normalise(header) -> str:
    if header is None:
        return ""
    return "".join(ch for ch in str(header).lower() if ch.isalnum())


def resolve_columns(header_row) -> dict[str, int]:
    """Map logical column names to 0-based indices by matching header text.

    Raises CheckError naming the missing column and listing what was found, so a
    layout mismatch fails loudly instead of silently reading the wrong column.
    """
    seen = {_normalise(cell): idx for idx, cell in enumerate(header_row)}
    resolved: dict[str, int] = {}

    for logical, aliases in {**COLUMN_ALIASES, **OPTIONAL_ALIASES}.items():
        for alias in aliases:
            if alias in seen:
                resolved[logical] = seen[alias]
                break

    missing = [name for name in REQUIRED_COLUMNS if name not in resolved]
    if missing:
        found = [str(h) for h in header_row if h is not None]
        raise CheckError(
            "could not find required column(s): " + ", ".join(missing)
            + ". Headers found in the sheet: " + (", ".join(found) or "(none)")
        )
    return resolved


def _is_percent_formatted(cell) -> bool:
    fmt = getattr(cell, "number_format", None)
    return bool(fmt) and "%" in str(fmt)


def read_rate(cell):
    """Return a rate as a percentage number (6.0 means 6%), or None if blank.

    A percent-formatted cell stores 0.06 for 6%, while a plain-number column may
    store 6.0 directly. Decide from the cell's number format rather than guessing
    from magnitude -- a genuine 0.5% rate would defeat any magnitude heuristic.
    """
    if cell is None or cell.value is None:
        return None
    value = cell.value
    if isinstance(value, str):
        value = value.strip().rstrip("%")
        if not value:
            return None
        try:
            return float(value)
        except ValueError as exc:
            raise CheckError(f"could not read {cell.coordinate!r} as a rate: {cell.value!r}") from exc
    if not isinstance(value, (int, float)):
        raise CheckError(f"could not read {cell.coordinate!r} as a rate: {value!r}")
    return float(value) * 100.0 if _is_percent_formatted(cell) else float(value)


def read_text(cell):
    if cell is None or cell.value is None:
        return None
    text = str(cell.value).strip()
    return text or None


def read_date(cell):
    """Return an ISO date string, or None if blank.

    Dates are written as text so they survive a round trip through either reader,
    but a sheet edited in Excel may come back as a datetime -- normalise both to
    YYYY-MM-DD so comparisons stay lexical.
    """
    if cell is None or cell.value is None:
        return None
    value = cell.value
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    return text[:10] if text else None


# --- Policy evaluation -------------------------------------------------------

def effective_rate(rate: float, override_code, override_rate) -> float:
    """What the customer is actually charged."""
    if not override_code or override_rate is None:
        return rate
    return min(rate, override_rate)


def _round2(value):
    return None if value is None else round(value + 0.0, 2)


def _finding(row, rule, branch, expected_value, actual_value, **values):
    """Build a violation.

    The reported figures are positional and the message placeholders arrive in
    `values`, so a template using {expected} or {actual} cannot collide with the
    parameters that carry those same figures into the report.
    """
    return {
        "rule": rule["id"],
        "severity": _severity(rule, branch),
        "accountId": row["accountId"],
        "rateType": row["rateType"],
        "scenario": row.get("scenario"),
        "expected": expected_value,
        "actual": _round2(actual_value),
        "message": _message(rule, branch, **values),
    }


def check_override_code_ceiling(row, rule, epsilon, violations):
    """A protected-population override code caps what may be charged.

    The cap is a ceiling, not a fixed rate: a customer whose normal formula lands
    below it keeps the lower rate (APR-099). Where a row carries several codes, the
    most protective ceiling governs (APR-162).
    """
    code = row["overrideCode"]
    if not code:
        return

    ceilings = rule["params"]["codeCeilings"]
    applicable = {c.strip().upper(): ceilings[c.strip().upper()]
                  for c in code.split(",")
                  if c.strip().upper() in ceilings}
    if not applicable:
        return

    winner = min(applicable, key=applicable.get)
    ceiling = applicable[winner]

    effective = row["effective"]
    if effective - ceiling <= epsilon:
        return

    violations.append(_finding(
        row, rule, "breach", ceiling, effective,
        actual=effective, ceiling=ceiling, code=winner))


def check_max_apr_cap(row, rule, epsilon, violations):
    """No card member may be charged above the ceiling their agreement discloses.

    Judged on the effective rate only: a normal rate above the cap that a lower
    override masks charges the customer the lower value, so nothing is overcharged
    and there is nothing to report.
    """
    max_rate = row.get("ceilingRate")
    if max_rate is None:
        max_rate = rule["params"].get("defaultMaxRate")
    if max_rate is None:
        return

    effective = row["effective"]
    if effective - max_rate <= epsilon:
        return

    violations.append(_finding(
        row, rule, "breach", max_rate, effective, actual=effective, max=max_rate))


def check_rate_floor(row, rule, epsilon, violations):
    """A disclosed floor must hold however far the index falls.

    Rows carrying an override are exempt: an override is a legitimate reason to sit
    below the floor, so applying this there would flag every protected account.
    """
    floor = row.get("floorRate")
    if floor is None or row["overrideCode"]:
        return

    effective = row["effective"]
    if floor - effective <= epsilon:
        return

    violations.append(_finding(
        row, rule, "breach", floor, effective, actual=effective, floor=floor))


def check_rate_matches_formula(row, rule, epsilon, violations):
    """A variable rate must equal index + margin at the disclosed precision.

    Checked against the normal rate rather than the effective one: an override
    changes what is charged, not how the underlying rate is derived.
    """
    basis = rule["params"].get("basis", "VARIABLE")
    if (row.get("rateBasis") or "").upper() != basis.upper():
        return

    index, margin = row.get("index"), row.get("margin")
    if index is None or margin is None:
        return

    derived = round(index + margin, rule["params"].get("precision", 2))
    rate = row["rate"]
    if abs(rate - derived) <= epsilon:
        return

    violations.append(_finding(
        row, rule, "drift", _round2(derived), rate,
        actual=rate, index=index, margin=margin, expected=derived))


def check_rate_sanity(row, rule, epsilon, violations):
    """A rate outside any plausible range is a data-integrity failure, not pricing."""
    effective = row["effective"]
    minimum = rule["params"].get("minRate", 0.0)
    absurd = rule["params"].get("absurdRate")

    if minimum - effective > epsilon:
        violations.append(_finding(
            row, rule, "negative", minimum, effective, actual=effective, min=minimum))
        return

    if absurd is not None and effective - absurd > epsilon:
        violations.append(_finding(
            row, rule, "absurd", absurd, effective, actual=effective, absurd=absurd))


def check_lower_rate_wins(row, rule, epsilon, violations):
    """Where an override exists, the lower of the two rates must be charged."""
    code = row["overrideCode"]
    if not code:
        return

    rate = row["rate"]
    override_rate = row["overrideRate"]

    if override_rate is None:
        violations.append(_finding(
            row, rule, "missingOverrideRate", None, rate, code=code, rate=rate))
        return

    # Regression guard: effective_rate() takes the minimum by construction, so
    # this cannot fire on this implementation. It exists to catch a system whose
    # served rate disagrees with the policy.
    expected = min(rate, override_rate)
    actual = row["effective"]
    if abs(actual - expected) > epsilon:
        violations.append(_finding(
            row, rule, "mismatch", _round2(expected), actual,
            actual=actual, rate=rate, overrideRate=override_rate, expected=expected))


def _as_of(rule):
    """The date the sheet is judged against. Pinned in the rule file so a report is
    reproducible; falling back to today would make yesterday's run unrepeatable."""
    return rule["params"].get("asOfDate") or datetime.now(timezone.utc).strftime("%Y-%m-%d")


def check_config_matches(row, rule, epsilon, violations):
    """A configuration column must match the value policy declares for the portfolio.

    Interest mechanics -- day-count divisor, compounding frequency, rounding rule --
    have to be uniform, because two accounts computing interest differently is a
    defect no matter which of them is 'right' (APR-041, APR-046, APR-048).
    """
    field = rule["params"]["field"]
    expected = rule["params"]["expected"]

    actual = row.get(field)
    if actual is None:
        return

    if str(actual).strip().upper() == str(expected).strip().upper():
        return

    # Template keys are want/got rather than expected/actual: _finding already takes
    # those two positionally, and reusing the names would collide in the call.
    violations.append(_finding(
        row, rule, "mismatch", expected, None,
        field=field, want=expected, got=actual))


def check_override_not_expired(row, rule, epsilon, violations):
    """An override past its expiry date should have reverted, not still be in force.

    A rate that sticks after the arrangement behind it ends is the APR-088 failure:
    the customer keeps a rate nobody re-authorised, and the underlying rate that
    should now apply is never recalculated.
    """
    if not row["overrideCode"]:
        return

    expiry = row.get("overrideExpiry")
    if not expiry:
        return

    as_of = _as_of(rule)
    if expiry >= as_of:
        return

    violations.append(_finding(
        row, rule, "expired", None, row["effective"],
        code=row["overrideCode"], expiry=expiry, asOf=as_of))


def check_protection_dates_valid(row, rule, epsilon, violations):
    """A protected-population cap needs a protection period behind it.

    Missing start dates mean the cap cannot be evidenced (APR-103); a cap still
    applied after the protection ended means the reversion never happened (APR-105).
    """
    code = row["overrideCode"]
    if not code:
        return

    protected = rule["params"]["protectedCodes"]
    applicable = [c.strip().upper() for c in code.split(",")
                  if c.strip().upper() in protected]
    if not applicable:
        return

    # A sheet without the protection columns cannot answer this; flagging every
    # protected row there would be an artefact of the layout, not a finding.
    if "protectionStart" not in row.get("_columnsPresent", frozenset()):
        return

    as_of = _as_of(rule)
    start, end = row.get("protectionStart"), row.get("protectionEnd")

    if not start:
        violations.append(_finding(
            row, rule, "missingStart", None, None, code=",".join(applicable)))
        return

    if end and end < as_of:
        violations.append(_finding(
            row, rule, "endedButApplied", None, row["effective"],
            code=",".join(applicable), end=end, asOf=as_of))


def check_pre_service_debt_scope(row, rule, epsilon, violations):
    """A statutory cap covers debt incurred before the protection began.

    Applying it to balances opened after activation extends the protection past what
    the statute grants, which is as much a control failure as under-applying it
    (APR-100).
    """
    code = row["overrideCode"]
    if not code:
        return

    protected = rule["params"]["protectedCodes"]
    applicable = [c.strip().upper() for c in code.split(",")
                  if c.strip().upper() in protected]
    if not applicable:
        return

    start, origin = row.get("protectionStart"), row.get("originationDate")
    if not start or not origin:
        return

    if origin < start:
        return

    violations.append(_finding(
        row, rule, "postActivation", None, row["effective"],
        code=",".join(applicable), origination=origin, start=start))


def check_bounds_within_product_range(row, rule, epsilon, violations):
    """An account's own floor and ceiling must sit inside its product's disclosed range.

    The marketed range has to bound every rate the system can actually assign, so an
    account configured outside it can be charged a rate that was never disclosed
    (APR-112, APR-073).
    """
    product = row.get("productCode")
    floor, ceiling = row.get("floorRate"), row.get("ceilingRate")
    if not product or floor is None or ceiling is None:
        return

    ranges = rule["params"]["productRanges"]
    bounds = ranges.get(product.strip().upper())
    if not bounds:
        violations.append(_finding(
            row, rule, "unknownProduct", None, None, product=product))
        return

    low, high = bounds
    if floor - low >= -epsilon and ceiling - high <= epsilon:
        return

    violations.append(_finding(
        row, rule, "outsideRange", None, None,
        product=product, floor=floor, ceiling=ceiling, low=low, high=high))


# Maps a rule file's "check" name to the function that evaluates it. Rule files
# select from these by name rather than carrying executable logic, so an untrusted
# rule file can retune the policy but cannot introduce new behaviour.
CHECKS = {
    "override_code_ceiling": check_override_code_ceiling,
    "max_apr_cap": check_max_apr_cap,
    "rate_floor": check_rate_floor,
    "lower_rate_wins": check_lower_rate_wins,
    "rate_matches_formula": check_rate_matches_formula,
    "rate_sanity": check_rate_sanity,
    "config_matches": check_config_matches,
    "override_not_expired": check_override_not_expired,
    "protection_dates_valid": check_protection_dates_valid,
    "pre_service_debt_scope": check_pre_service_debt_scope,
    "bounds_within_product_range": check_bounds_within_product_range,
}


# --- Reading and reporting ---------------------------------------------------

def load_rows(path: str, sheet_name: str | None):
    try:
        workbook = load_workbook(path, data_only=True)
    except FileNotFoundError as exc:
        raise CheckError(f"file not found: {path}") from exc
    except Exception as exc:
        raise CheckError(f"could not open {path}: {exc}") from exc

    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
    except KeyError as exc:
        available = ", ".join(workbook.sheetnames)
        raise CheckError(f"no sheet named {sheet_name!r}. Available: {available}") from exc

    rows = list(worksheet.iter_rows())
    if not rows:
        raise CheckError(f"sheet {worksheet.title!r} is empty")

    columns = resolve_columns([cell.value for cell in rows[0]])

    def cell_at(cells, name):
        idx = columns.get(name)
        if idx is None or idx >= len(cells):
            return None
        return cells[idx]

    records = []
    for cells in rows[1:]:
        account_id = read_text(cells[columns["accountId"]]) if columns["accountId"] < len(cells) else None
        if account_id is None:
            continue  # trailing blank row

        rate = read_rate(cells[columns["rate"]]) or 0.0
        override_code = read_text(cells[columns["overrideCode"]])
        override_rate = read_rate(cells[columns["overrideRate"]])

        records.append({
            "accountId": account_id,
            "rateType": read_text(cells[columns["rateType"]]),
            "rate": rate,
            "overrideCode": override_code,
            "overrideRate": override_rate,
            "effective": effective_rate(rate, override_code, override_rate),
            # Absent when the sheet does not carry the column; rules needing them skip.
            "scenario": read_text(cell_at(cells, "scenario")),
            "rateBasis": read_text(cell_at(cells, "rateBasis")),
            "index": read_rate(cell_at(cells, "index")),
            "margin": read_rate(cell_at(cells, "margin")),
            "floorRate": read_rate(cell_at(cells, "floorRate")),
            "ceilingRate": read_rate(cell_at(cells, "ceilingRate")),
            "productCode": read_text(cell_at(cells, "productCode")),
            "dayCountBasis": read_text(cell_at(cells, "dayCountBasis")),
            "compoundingFrequency": read_text(cell_at(cells, "compoundingFrequency")),
            "roundingRule": read_text(cell_at(cells, "roundingRule")),
            **{name: read_date(cell_at(cells, name)) for name in DATE_COLUMNS},
            # Which optional columns the sheet actually carries. A rule must be able
            # to tell "column absent" from "cell blank": the first means the sheet
            # cannot answer the question, the second is a real gap in the data.
            "_columnsPresent": frozenset(columns),
        })
    return records


def build_report(records, ruleset=None) -> dict:
    """Evaluate every record against every enabled rule.

    ruleset defaults to the bundled rules.json, so callers that do not care which
    policy is in force can keep calling this with just the records.
    """
    ruleset = ruleset if ruleset is not None else load_rules()
    epsilon = ruleset.get("epsilon", EPSILON)
    rules = active_rules(ruleset)

    violations: list[dict] = []
    for row in records:
        # Rule order comes from the file: it is the order findings are presented
        # in, and matches how the rules are numbered in SKILL.md.
        for rule in rules:
            CHECKS[rule["check"]](row, rule, epsilon, violations)

    critical = sum(1 for v in violations if v["severity"] == SEVERITY_CRITICAL)

    return {
        "status": "compliant" if not violations else "violations_found",
        "checkedAt": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "rowsChecked": len(records),
        "accountsChecked": len({r["accountId"] for r in records}),
        "criticalCount": critical,
        "highCount": len(violations) - critical,
        "violations": violations,
        "compliant": not violations,
        "rulesVersion": ruleset.get("version"),
        "rulesApplied": [rule["id"] for rule in rules],
    }


def render_text(report: dict) -> str:
    lines = []
    if report["status"] == "compliant":
        lines.append(f"COMPLIANT - {report['rowsChecked']} rows across "
                     f"{report['accountsChecked']} accounts, no policy violations.")
        return "\n".join(lines)

    lines.append(f"NON-COMPLIANT - {len(report['violations'])} violation(s): "
                 f"{report['criticalCount']} critical, {report['highCount']} high "
                 f"(across {report['rowsChecked']} rows, "
                 f"{report['accountsChecked']} accounts).")
    lines.append("")
    width = max((len(v["accountId"]) for v in report["violations"]), default=9)
    for v in report["violations"]:
        expected = "-" if v["expected"] is None else f"{v['expected']:.2f}%"
        actual = "-" if v["actual"] is None else f"{v['actual']:.2f}%"
        lines.append(f"  {v['severity']:<8} {v['accountId']:<{width}} "
                     f"rate type {v['rateType']:<4} {v['rule']}")
        lines.append(f"           expected {expected}, actual {actual}")
        lines.append(f"           {v['message']}")
        lines.append("")
    return "\n".join(lines).rstrip()


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(
        description="Check an APR rate spreadsheet against company rate policy.")
    parser.add_argument("--file", required=True,
                        help="path to the .xlsx workbook to check")
    parser.add_argument("--sheet", default=None,
                        help="sheet name (default: the first sheet)")
    parser.add_argument("--format", choices=("json", "text"), default="json",
                        help="output format (default: json)")
    parser.add_argument("--rules", default=None,
                        help=f"path to the rule file (default: {DEFAULT_RULES_FILE})")
    args = parser.parse_args(argv)

    try:
        ruleset = load_rules(args.rules)
        records = load_rows(args.file, args.sheet)
        report = build_report(records, ruleset)
    except CheckError as exc:
        print(f"error: {exc}", file=sys.stderr)
        return EXIT_ERROR

    if args.format == "json":
        print(json.dumps(report, indent=2))
    else:
        print(render_text(report))

    return EXIT_COMPLIANT if report["compliant"] else EXIT_VIOLATIONS


if __name__ == "__main__":
    sys.exit(main())
